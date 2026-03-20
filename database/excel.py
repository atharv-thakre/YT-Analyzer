import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from database.db_pipeline import get_playlist_stats, get_playlist_videos


# =========================
# ⏱ Helper: seconds → HH:MM:SS
# =========================
def seconds_to_hms(seconds):
    if not seconds:
        return "00:00:00"

    seconds = int(seconds)
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60

    return f"{h:02d}:{m:02d}:{s:02d}"


# =========================
# 📊 Pretty Labels
# =========================
PRETTY_STATS = {
    "total_videos": "Total Videos",
    "avg_views": "Average Views",
    "avg_likes": "Average Likes",
    "avg_comments": "Average Comments",
    "avg_duration_sec": "Average Duration",
    "total_duration_sec": "Total Duration",
    "engagement_avg": "Average Engagement",
}


def export_playlist_report(session, playlist_id: str):

    videos = get_playlist_videos(session, playlist_id)
    stats = get_playlist_stats(session, playlist_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Playlist Report"

    # =========================
    # 🏆 TITLE
    # =========================
    ws.merge_cells("A1:G1")
    ws["A1"] = "PLAYLIST ANALYTICS REPORT"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    # =========================
    # 📌 PLAYLIST INFO
    # =========================
    ws["A3"] = "Playlist:"
    ws["B3"] = stats["playlist_title"]

    ws["A4"] = "Channel:"
    ws["B4"] = stats["playlist_channel"]

    ws["A5"] = "Playlist ID:"
    ws["B5"] = stats["playlist_id"]

    for r in range(3, 6):
        ws[f"A{r}"].font = Font(bold=True, size=12)
        ws[f"B{r}"].font = Font(size=12)

    # =========================
    # 📊 STATS
    # =========================
    ws["A7"] = "STATS"
    ws["A7"].font = Font(bold=True, size=13)
    ws.freeze_panes = None

    row = 9

    for key, value in stats.items():
        if key in ["playlist_id", "playlist_title", "playlist_channel"]:
            continue  # skip meta

        # Format duration
        if key in ["total_duration_sec", "avg_duration_sec"]:
            value = seconds_to_hms(value)

        # Round floats
        if isinstance(value, float):
            value = round(value, 4)

        label = PRETTY_STATS.get(key, key)

        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(bold=True)

        ws.cell(row=row, column=2, value=value)

        row += 1

    # =========================
    # 📄 VIDEOS TABLE
    # =========================
    row += 3

    ws.cell(row=row, column=1, value="VIDEOS")
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)

    row += 2

    headers = [
        "Position",
        "Title",
        "Channel",
        "Published At",
        "Views",
        "Likes",
        "Comments",
    ]

    header_row = row

    # Header
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, size=11)

    row += 1

    # Data rows
    data_start_row = row

    for v in videos:
        ws.cell(row=row, column=1, value=v["position"])
        ws.cell(row=row, column=2, value=v["title"])
        ws.cell(row=row, column=3, value=v["channel_name"])
        ws.cell(
            row=row,
            column=4,
            value=v["published_at"].strftime("%Y-%m-%d")
            if v["published_at"] else None
        )
        ws.cell(row=row, column=5, value=v["views"])
        ws.cell(row=row, column=6, value=v["likes"])
        ws.cell(row=row, column=7, value=v["comments"])

        row += 1

    data_end_row = row - 1

    # =========================
    # 📐 FORMATTING
    # =========================

    # Auto width ONLY for used columns
    for col in range(1, 8):
        max_length = 0
        col_letter = get_column_letter(col)

        for r in range(1, data_end_row + 1):
            val = ws.cell(row=r, column=col).value
            if val:
                max_length = max(max_length, len(str(val)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # Enable filter only on actual data
    ws.auto_filter.ref = f"A{header_row}:G{data_end_row}"

    # Save
    file_path = export_path(playlist_id)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)

    return file_path



def export_path(playlist_id: str) -> str:
    return f"sheets/playlist_report_{playlist_id}.xlsx"
